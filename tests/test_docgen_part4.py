# pyright: reportUnknownVariableType=false, reportUnknownMemberType=false, reportUnknownParameterType=false, reportMissingParameterType=false, reportUnknownArgumentType=false, reportPrivateUsage=false
# ruff: noqa: F403, F405
from ._docgen_fixture_builders import *
from ._docgen_test_support import *


def test_configgen_collect_and_process_components_cover_section_routing(tmp_path):
    class StubExtractor:
        unitlib_dir = Path("unitlib")
        projectlib_dir = Path("projectlib")
        nnelib_dir = Path("nnelib")
        sglib_dir = Path("SL_Library")

        def get_z_files(self, directory):
            mapping = {
                self.unitlib_dir: [Path("PBSLC100.z"), Path("IgnoreMe.z")],
                self.projectlib_dir: [Path("ProjectA.z")],
                self.nnelib_dir: [],
                self.sglib_dir: [Path("Support.z")],
            }
            return mapping[directory]

        def get_ip_address(self, z_file):
            return {"PBSLC100.z": "10.0.0.1", "IgnoreMe.z": "No Q-File"}[z_file.name]

        def get_component_info(self, z_file, component_type, has_ip, slc_programs):
            return configgen.ComponentInfo(
                name=z_file.stem,
                type=component_type,
                ip_address=slc_programs.get("PBSLC100", "N/A") if has_ip else "N/A",
                slc="SLC100" if z_file.stem == "PBSLC100" else "N/A",
                units="(1) UnitA" if has_ip else "N/A",
                dependencies=["Dep1"] if z_file.stem != "Support" else [],
            )

    generator = configgen.ExcelGenerator(extractor=_typed_extractor(StubExtractor()))

    slc_programs = generator._collect_slc_programs()
    component_data, dependency_data = generator._process_all_components(slc_programs)

    assert slc_programs == {"PBSLC100": "10.0.0.1"}
    assert [(component.name, component.type) for component in component_data] == [
        ("PBSLC100", "Program"),
        ("IgnoreMe", "Program"),
        ("ProjectA", "Project Library"),
        ("Support", "SG Library"),
    ]
    assert dependency_data == [("PBSLC100", "Dep1"), ("IgnoreMe", "Dep1"), ("ProjectA", "Dep1")]


def test_configgen_generate_runs_full_workbook_pipeline(tmp_path):
    output_path = tmp_path / "config.xlsx"

    class StubExtractor:
        unitlib_dir = Path("unitlib")
        projectlib_dir = Path("projectlib")
        nnelib_dir = Path("nnelib")
        sglib_dir = Path("SL_Library")

        def get_z_files(self, directory):
            mapping = {
                self.unitlib_dir: [Path("PBSLC100.z")],
                self.projectlib_dir: [Path("CommonLib.z")],
                self.nnelib_dir: [],
                self.sglib_dir: [],
            }
            return mapping[directory]

        def get_ip_address(self, z_file):
            return "10.0.0.1" if z_file.stem == "PBSLC100" else "No Q-File"

        def get_component_info(self, z_file, component_type, has_ip, slc_programs):
            return configgen.ComponentInfo(
                name=z_file.stem,
                type=component_type,
                ip_address=slc_programs.get(z_file.stem, "N/A") if has_ip else "N/A",
                slc="SLC100" if z_file.stem == "PBSLC100" else "N/A",
                units="(1) UnitA" if has_ip else "N/A",
                dependencies=["CommonLib"] if z_file.stem == "PBSLC100" else [],
            )

        def parse_all_configuration_files(self):
            return [
                configgen.ConfigurationFileInfo(
                    config_name="KaGC_Allf",
                    version="1.0",
                    date="2026-04-30",
                    main_program="PBSLC100",
                    programs=[{"name": "PBSLC100", "directory": "unitlib", "main_program": True}],
                    libraries=[{"name": "CommonLib", "directory": "projectlib"}],
                )
            ]

    generator = configgen.ExcelGenerator(extractor=_typed_extractor(StubExtractor()))
    generator.workstation_mapper.workstation_map = {"KaGC_Allf": ["OP11"]}
    generator.workstation_mapper.physical_locations = {"OP11": "Control Room 3"}

    generator.generate(output_path)

    assert output_path.exists()
    assert {sheet.title for sheet in generator.wb.worksheets} == {
        "Query Tool",
        "Dashboard",
        "System Components",
        "Library Dependencies",
        "Station Configuration",
        "Configuration Summary",
        "Configuration Details",
    }
    assert generator.components_ws["B2"].value == "PBSLC100"
    assert generator.dependencies_ws["D2"].value == "CommonLib"


def test_configgen_main_returns_2_for_invalid_root(tmp_path, capsys):
    invalid_root = tmp_path / "missing-root"

    result = configgen.main([str(invalid_root)])

    captured = capsys.readouterr()
    assert result == 2
    assert "Invalid root directory" in captured.err


def test_configgen_main_generates_output_with_stubbed_components(tmp_path, monkeypatch, capsys):
    root_dir = tmp_path / "root"
    root_dir.mkdir()
    output_path = tmp_path / "out.xlsx"
    captured = {}

    class StubExtractor:
        def __init__(self, root):
            captured["extractor_root"] = root

    class StubGenerator:
        def __init__(self, extractor):
            captured["generator_extractor"] = extractor

        def generate(self, output):
            captured["output"] = output

    monkeypatch.setattr(configgen, "SattLineConfigExtractor", StubExtractor)
    monkeypatch.setattr(configgen, "ExcelGenerator", StubGenerator)

    result = configgen.main([str(root_dir), "--output", str(output_path)])

    captured_output = capsys.readouterr().out
    assert result == 0
    assert captured["extractor_root"] == root_dir.resolve()
    assert captured["output"] == output_path.resolve()
    assert "Configuration Excel file generated successfully" in captured_output


def test_configgen_main_returns_1_when_generation_fails(tmp_path, monkeypatch, capsys):
    root_dir = tmp_path / "root"
    root_dir.mkdir()

    class StubExtractor:
        def __init__(self, root):
            self.root = root

    class StubGenerator:
        def __init__(self, extractor):
            self.extractor = extractor

        def generate(self, _output):
            raise RuntimeError("generation failed")

    monkeypatch.setattr(configgen, "SattLineConfigExtractor", StubExtractor)
    monkeypatch.setattr(configgen, "ExcelGenerator", StubGenerator)

    result = configgen.main([str(root_dir)])

    captured = capsys.readouterr()
    assert result == 1
    assert "Error: generation failed" in captured.out


def test_configgen_main_module_exits_with_main_return_code(monkeypatch, tmp_path):
    missing_root = tmp_path / "missing-root"
    monkeypatch.setattr("sys.argv", ["configgen.py", str(missing_root)])

    with pytest.raises(SystemExit) as exc:
        runpy.run_module("sattlint.docgenerator.configgen", run_name="__main__")

    assert exc.value.code == 2


def test_configgen_style_manager_applies_header_styling():
    from openpyxl import Workbook  # noqa: PLC0415

    wb = Workbook()
    ws = _active_worksheet(wb)

    style_manager = configgen.StyleManager()
    cell = ws["A1"]
    cell.value = "Test Header"
    style_manager.apply_header_style(cell)

    assert cell.font.bold is True
    assert cell.font.size == 11
    assert "4472C4" in cell.fill.start_color.rgb


def test_configgen_worksheet_helper_setup_headers():
    from openpyxl import Workbook  # noqa: PLC0415

    wb = Workbook()
    ws = _active_worksheet(wb)
    style_manager = configgen.StyleManager()
    headers = ["ID", "Name", "Value"]

    configgen.WorksheetHelper.setup_headers(ws, headers, style_manager)

    assert ws["A1"].value == "ID"
    assert ws["B1"].value == "Name"
    assert ws["C1"].value == "Value"
    assert ws["A1"].font.bold is True


def test_configgen_worksheet_helper_create_table():
    from openpyxl import Workbook  # noqa: PLC0415

    wb = Workbook()
    ws = _active_worksheet(wb)

    configgen.WorksheetHelper.create_table(ws, "TestTable", "A1:C3")

    assert len(ws.tables) == 1
    assert ws.tables["TestTable"] is not None


def test_configgen_worksheet_helper_auto_fit_columns():
    from openpyxl import Workbook  # noqa: PLC0415

    wb = Workbook()
    ws = _active_worksheet(wb)
    ws["A1"] = "Short"
    ws["B1"] = "This is a much longer value that should increase column width"

    configgen.WorksheetHelper.auto_fit_columns(ws, max_width=60)

    assert ws.column_dimensions["A"].width > 0
    assert ws.column_dimensions["B"].width >= ws.column_dimensions["A"].width


def test_configgen_aggregate_slc_numbers_matches_component_case_insensitive():
    component_data = [
        configgen.ComponentInfo(
            name="PBSLC123", type="Program", ip_address="10.0.0.1", slc="SLC123", units="(2) A, B", dependencies=[]
        ),
        configgen.ComponentInfo(
            name="OpsStation", type="Program", ip_address="10.0.0.2", slc="No SLC", units="(0) ", dependencies=[]
        ),
    ]
    config = {"programs": ["pbslc123", "OpsStation"]}
    generator = configgen.ExcelGenerator.__new__(configgen.ExcelGenerator)

    result = generator._aggregate_slc_numbers(config, component_data)

    assert result == "SLC123"


def test_configgen_aggregate_units_extracts_from_format():
    component_data = [
        configgen.ComponentInfo(
            name="Prog1",
            type="Program",
            ip_address="10.0.0.1",
            slc="SLC123",
            units="(3) UnitA, UnitB, UnitC",
            dependencies=[],
        ),
        configgen.ComponentInfo(
            name="Prog2",
            type="Program",
            ip_address="10.0.0.2",
            slc="No SLC",
            units="(1) UnitD",
            dependencies=[],
        ),
    ]
    config = {"programs": ["Prog1", "Prog2"]}
    generator = configgen.ExcelGenerator.__new__(configgen.ExcelGenerator)

    result = generator._aggregate_units(config, component_data)

    assert "UnitA" in result
    assert "UnitB" in result
    assert "UnitC" in result
    assert "UnitD" in result
    assert result.startswith("(")


def test_configgen_format_units_for_station_returns_no_units_for_invalid_input():
    generator = configgen.ExcelGenerator.__new__(configgen.ExcelGenerator)

    assert generator._format_units_for_station("") == "No units assigned"
    assert generator._format_units_for_station("No X-File") == "No units assigned"
    assert generator._format_units_for_station("Error reading X-File") == "No units assigned"


def test_configgen_format_units_for_station_extracts_from_paren_format():
    generator = configgen.ExcelGenerator.__new__(configgen.ExcelGenerator)

    result = generator._format_units_for_station("(3) UnitA, UnitB, UnitC")

    assert result == "UnitA, UnitB, UnitC"


def test_configgen_format_units_for_station_returns_original_text_for_unmatched_paren_format():
    generator = configgen.ExcelGenerator.__new__(configgen.ExcelGenerator)

    result = generator._format_units_for_station("(3)UnitA, UnitB")

    assert result == "(3)UnitA, UnitB"


def test_configgen_determine_station_type_classifies_by_prefix():
    generator = configgen.ExcelGenerator.__new__(configgen.ExcelGenerator)

    assert generator._determine_station_type("LOP01") == "Local Operator Panel"
    assert generator._determine_station_type("OP06") == "Operator Station"
    assert generator._determine_station_type("OPC01") == "OPC Server"
    assert generator._determine_station_type("PRG01") == "Programmer Station"
    assert generator._determine_station_type("Journal Server 1 Primary") == "Journal Server"
    assert generator._determine_station_type("UNKNOWN") == "Special System"


def test_configgen_populate_components_sheet_writes_rows_and_returns_count():
    generator = configgen.ExcelGenerator(extractor=_typed_extractor(object()))
    component_data = [
        configgen.ComponentInfo(
            name="PBSLC123",
            type="Program",
            ip_address="10.0.0.1",
            slc="SLC123",
            units="(2) UnitA, UnitB",
            dependencies=["CommonLib"],
        ),
        configgen.ComponentInfo(
            name="CommonLib",
            type="Project Library",
            ip_address="N/A",
            slc="No SLC",
            units="No X-File",
            dependencies=[],
        ),
    ]

    row_count = generator._populate_components_sheet(component_data)

    assert row_count == 3
    assert generator.components_ws["A1"].value == "ID"
    assert generator.components_ws["B2"].value == "PBSLC123"
    assert generator.components_ws["F3"].value == "Project Library"


def test_configgen_populate_dependencies_sheet_uses_case_insensitive_type_lookup():
    generator = configgen.ExcelGenerator(extractor=_typed_extractor(object()))
    generator.all_component_data = [
        configgen.ComponentInfo(
            name="PBSLC123",
            type="Program",
            ip_address="10.0.0.1",
            slc="SLC123",
            units="(1) UnitA",
            dependencies=["CommonLib"],
        ),
        configgen.ComponentInfo(
            name="CommonLib",
            type="Project Library",
            ip_address="N/A",
            slc="No SLC",
            units="No X-File",
            dependencies=[],
        ),
    ]

    row_count = generator._populate_dependencies_sheet([("pbslc123", "commonlib"), ("UnknownComp", "UnknownLib")])

    assert row_count == 3
    assert generator.dependencies_ws["C2"].value == "Program"
    assert generator.dependencies_ws["E2"].value == "Project Library"
    assert generator.dependencies_ws["C3"].value == "Unknown"
    assert generator.dependencies_ws["E3"].value == "Unknown"
