"""
sattline_config_excel.py

Generates an Excel spreadsheet documenting SattLine program configurations
in normalized long format with Excel Tables and an interactive Dashboard.
"""

from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import logging
from dataclasses import dataclass
from functools import lru_cache
from typing import Optional

log = logging.getLogger("SattLineConfigExcel")


@dataclass
class ComponentInfo:
    """Data class for component information."""
    name: str  # WITHOUT .z extension, ORIGINAL CASE preserved
    type: str
    ip_address: str
    slc: str
    units: str
    dependencies: list[str]  # WITHOUT .z extensions, ORIGINAL CASE preserved


@dataclass
class ConfigurationFileInfo:
    """Data class for configuration file information."""
    config_name: str  # ORIGINAL CASE preserved
    version: str
    date: str
    main_program: str  # ORIGINAL CASE preserved
    programs: list[dict]  # names in ORIGINAL CASE
    libraries: list[dict]  # names in ORIGINAL CASE


class ExcelConfig:
    """Configuration constants for Excel generation."""
    HEADER_COLOR = "4472C4"
    HEADER_TEXT_COLOR = "FFFFFF"
    KPI_TITLE_COLOR = "1F4E78"
    KPI_VALUE_COLOR = "2E75B6"
    KPI_BG_COLOR = "E7E6E6"
    SUCCESS_COLOR = "70AD47"
    WARNING_COLOR = "ED7D31"


class ConfigurationFileParser:
    """Parser for SattLine configuration files (.k files)."""
    
    def __init__(self):
        self.config_pattern = re.compile(
            r'Configuration\s*\(\s*Version\s+"([^"]+)"\s+Date\s+"([^"]+)"\s+Name\s+"([^"]+)"',
            re.DOTALL
        )
        self.program_pattern = re.compile(
            r'Program\s*\(\s*Name\s+"([^"]+)"\s+Directory\s+"([^"]+)"\s+MainProgram\s+(\w+)',
            re.DOTALL
        )
        self.library_pattern = re.compile(
            r'Library\s*\(\s*Name\s+"([^"]+)"\s+Directory\s+"([^"]+)"',
            re.DOTALL
        )
    
    def parse_configuration_file(self, config_file: Path) -> Optional[ConfigurationFileInfo]:
        """Parse a configuration file and extract all programs and libraries."""
        try:
            text = read_text_with_fallback(config_file)
            
            # Extract configuration header
            config_match = self.config_pattern.search(text)
            if not config_match:
                log.warning(f"Could not parse configuration header in {config_file.name}")
                return None
            
            version = config_match.group(1)
            date = config_match.group(2)
            # Use the filename (without .k extension) as the configuration name
            config_name = config_file.stem  # This gets filename without extension
            
            # Extract all programs (remove .z extension if present, keep original case)
            programs = []
            for match in self.program_pattern.finditer(text):
                program_name_raw = match.group(1)
                # Only remove .z extension if it exists
                if program_name_raw.endswith('.z'):
                    program_name = program_name_raw[:-2]
                else:
                    program_name = program_name_raw
                
                directory = match.group(2)
                main_program = match.group(3) == 'True'
                
                programs.append({
                    'name': program_name,
                    'directory': directory,
                    'main_program': main_program
                })
            
            # Extract all libraries (remove .z extension if present, keep original case)
            libraries = []
            for match in self.library_pattern.finditer(text):
                library_name_raw = match.group(1)
                # Only remove .z extension if it exists
                if library_name_raw.endswith('.z'):
                    library_name = library_name_raw[:-2]
                else:
                    library_name = library_name_raw
                
                directory = match.group(2)
                
                libraries.append({
                    'name': library_name,
                    'directory': directory
                })
            
            # Find main program
            main_program = next((p['name'] for p in programs if p['main_program']), "None")
            
            log.info(f"✓ Parsed {config_file.name}: Config='{config_name}', {len(programs)} programs, {len(libraries)} libraries")
            
            return ConfigurationFileInfo(
                config_name=config_name,
                version=version,
                date=date,
                main_program=main_program,
                programs=programs,
                libraries=libraries
            )
            
        except Exception as e:
            log.error(f"Error parsing configuration file {config_file}: {e}")
            return None


class WorkstationMapper:
    """Maps SattLine configurations to physical workstations."""
    
    def __init__(self):
        # Map configuration files (without .z) to workstations with physical locations
        # Keys are stored in original case for display
        self.workstation_map = {
            "KaGC_OP_Utilf": ["LOP01", "OP06", "OP07"],
            "KaGC_OP_Procesf": ["OP01", "OP02", "OP03", "OP04", "OP05", "OP10"],
            "KaGC_Allf": ["OP11", "OP12", "LOP14"],
            "KaGC_Autf": ["PRG01", "PRG02"],
            "KaGC_OP_FDf": ["LOP11", "LOP12", "LOP13"],
            "KaGC_LOP04": ["LOP04"],
            "KaGC_LOP05": ["LOP05"],
            "KaGC_LOP08": ["LOP08"],
            "KaGC_LOP09": ["LOP09"],
            "KaGC_LOP10": ["LOP10"],
            "KaGC_LOP17": ["LOP17"],
            "KaGC_LOP18": ["LOP18"],
            "KaGC_LOP19": ["LOP19"],
            "KaGC_OPC01f": ["OPC01"],
            "KaGC_OPC02f": ["OPC02"],
            "KaGC_OPC03f": ["OPC03"],
            "KaGC_OPC04f": ["OPC04"],
            "KaGC_OPC05f": ["OPC05"],
            "KaGC_OPC06f": ["OPC06"],
            "KaGC_OPC07f": ["OPC07"],
            "KaGC_OPC08f": ["OPC08"],
            "KaGC_OPC09f": ["OPC09"],
            "KaGC_OPC10f": ["OPC10"],
            "KaGC_OPC11f": ["OPC11"],
            "SGDKKAGC01f": ["Kurver"],
            "KaGC_JN01f": ["Journal Server 1 Primær"],
            "KaGC_JN02f": ["Journal Server 1 Sekundær"],
            "KaGC_JN0304f": ["Journal Server 2 Primær", "Journal Server 2 Sekundær"],
        }
        
        # Physical location mapping
        self.physical_locations = {
            "LOP01": "Control Room 1",
            "OP06": "Control Room 2",
            "OP07": "Control Room 2",
            "OP01": "Main Control Room",
            "OP02": "Main Control Room",
            "OP03": "Main Control Room",
            "OP04": "Main Control Room",
            "OP05": "Main Control Room",
            "OP10": "Main Control Room",
            "OP11": "Control Room 3",
            "OP12": "Control Room 3",
            "LOP14": "Local Panel Area",
            "PRG01": "Engineering Office",
            "PRG02": "Engineering Office",
            "LOP11": "Field Station Area 1",
            "LOP12": "Field Station Area 1",
            "LOP13": "Field Station Area 1",
            "LOP04": "Field Station Area 2",
            "LOP05": "Field Station Area 2",
            "LOP08": "Field Station Area 3",
            "LOP09": "Field Station Area 3",
            "LOP10": "Field Station Area 3",
            "LOP17": "Field Station Area 4",
            "LOP18": "Field Station Area 4",
            "LOP19": "Field Station Area 4",
            "OPC01": "Server Room",
            "OPC02": "Server Room",
            "OPC03": "Server Room",
            "OPC04": "Server Room",
            "OPC05": "Server Room",
            "OPC06": "Server Room",
            "OPC07": "Server Room",
            "OPC08": "Server Room",
            "OPC09": "Server Room",
            "OPC10": "Server Room",
            "OPC11": "Server Room",
            "Kurver": "Curve Management Room",
            "Journal Server 1 Primær": "Server Room",
            "Journal Server 1 Sekundær": "Server Room",
            "Journal Server 2 Primær": "Server Room",
            "Journal Server 2 Sekundær": "Server Room",
        }
    
    def get_workstations(self, component_name: str) -> list[str]:
        """Get workstation(s) for a given component (without extension, case-insensitive lookup)."""
        component_name = component_name.replace('.z', '')
        # Case-insensitive lookup
        for key in self.workstation_map:
            if key.lower() == component_name.lower():
                return self.workstation_map[key]
        return []
    
    def get_physical_location(self, station_id: str) -> str:
        """Get physical location for a workstation."""
        return self.physical_locations.get(station_id, "Unknown Location")


def read_text_with_fallback(file_path: Path) -> str:
    """Read text file with utf-8, falling back to cp1252 for Danish characters."""
    try:
        return file_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        log.debug(f"UTF-8 decode failed for {file_path.name}, trying cp1252")
        return file_path.read_text(encoding="cp1252")


class SattLineConfigExtractor:
    """Extracts configuration information from SattLine project files."""
    
    def __init__(self, root_dir: Path):
        self.root_dir = root_dir
        self.unitlib_dir = root_dir / "unitlib"
        self.projectlib_dir = root_dir / "projectlib"
        self.nnelib_dir = root_dir / "nnelib"
        self.sglib_dir = root_dir / "SL_Library"
        self.kfiles_dir = root_dir / "Configuration"
        
        # Compiled regex patterns
        self.slc_pattern = re.compile(r'SLC(\d+)', re.IGNORECASE)
        self.pbslc_pattern = re.compile(r'PBSLC(\d+)', re.IGNORECASE)
        self.name_pattern = re.compile(r'\(\s*Name\s+"([^"]+)"')
        self.unit_pattern = re.compile(r'\bp(\w+)\s*(?:"[^"]*")?\s*:\s*pType\s*;')
        
        if not self.validate_directories():
            raise ValueError("Required directories not found")

    def validate_directories(self) -> bool:
        """Validate that all required directories exist."""
        required_dirs = [
            self.unitlib_dir,
            self.projectlib_dir,
            self.nnelib_dir,
            self.sglib_dir
        ]
        
        missing_dirs = [d for d in required_dirs if not d.exists()]
        
        if missing_dirs:
            log.error(f"Missing directories: {[str(d) for d in missing_dirs]}")
            return False
        
        log.info("✓ All required directories found")
        return True
    
    def parse_all_configuration_files(self) -> list[ConfigurationFileInfo]:
        """Parse all configuration files (.k files)."""
        parser = ConfigurationFileParser()
        configurations = []
        
        if not self.kfiles_dir.exists():
            log.warning(f"Configuration directory not found: {self.kfiles_dir}")
            return configurations
        
        config_files = sorted(self.kfiles_dir.glob("*.k"))
        
        if not config_files:
            log.warning(f"No .k files found in {self.kfiles_dir}")
            return configurations
        
        log.info(f"📁 Found {len(config_files)} configuration files")
        
        for config_file in config_files:
            config_info = parser.parse_configuration_file(config_file)
            if config_info:
                configurations.append(config_info)
        
        log.info(f"✓ Successfully parsed {len(configurations)} configuration files")
        return configurations
    
    def get_z_files(self, directory: Path) -> list[Path]:
        """Get all .z files from a directory."""
        if not directory.exists():
            log.warning(f"Directory not found: {directory}")
            return []
        return sorted(directory.glob("*.z"))
    
    @lru_cache(maxsize=512)
    def _read_file_cached(self, file_path: Path) -> str:
        """Cache file reads to avoid redundant I/O."""
        return read_text_with_fallback(file_path)
    
    def read_dependencies(self, z_file: Path) -> list[str]:
        """Read dependency list from .z file and remove .z extensions (preserve original case)."""
        try:
            text = self._read_file_cached(z_file)
            # Remove .z extension from each dependency, preserve original case
            deps = [line.strip().replace('.z', '') for line in text.splitlines() if line.strip()]
            return deps
        except Exception as e:
            log.error(f"Error reading {z_file}: {e}")
            return []
    
    def get_ip_address(self, z_file: Path) -> str:
        """Extract IP address from corresponding .q file."""
        q_file = z_file.with_suffix(".q")
        if not q_file.exists():
            return "No Q-File"
        
        try:
            text = read_text_with_fallback(q_file)
            match = self.name_pattern.search(text)
            return match.group(1) if match else "No SLC assigned"
        except Exception as e:
            log.error(f"Error reading {q_file}: {e}")
            return "Error reading Q-File"
    
    def get_slc_name(self, ip_address: str, slc_programs: dict[str, str]) -> str:
        """Find SLC name by matching IP address (case-insensitive)."""
        if ip_address in ["No Q-File", "No SLC assigned", "Error reading Q-File"]:
            return "No SLC"
        
        for prog_name, prog_ip in slc_programs.items():
            if prog_ip == ip_address:
                if "pbslc" in prog_name.lower():
                    if match := self.pbslc_pattern.search(prog_name):
                        return f"SLC{match.group(1)}"
                elif "wdslc" in prog_name.lower() or prog_name.lower().startswith("kagcwd"):
                    if match := self.slc_pattern.search(prog_name):
                        return f"SLC{match.group(1)}"
        
        return "No SLC"
    
    def get_units_in_program(self, z_file: Path) -> str:
        """Extract unit names from corresponding .x file."""
        x_file = z_file.with_suffix(".x")
        if not x_file.exists():
            return "No X-File"
        
        try:
            text = read_text_with_fallback(x_file)
            units_set = {match.group(1) for match in self.unit_pattern.finditer(text)}
            
            if not units_set:
                return "No units assigned"
            
            units = sorted(units_set)
            return f"({len(units)}) " + ", ".join(units)
        
        except Exception as e:
            log.error(f"Error reading {x_file}: {e}")
            return "Error reading X-File"
    
    def get_component_info(self, z_file: Path, component_type: str, 
                          has_ip: bool, slc_programs: dict[str, str]) -> ComponentInfo:
        """Extract all component information (WITHOUT .z extension, preserve original case)."""
        # Remove .z extension from component name, preserve original case
        component_name = z_file.stem
        dependencies = self.read_dependencies(z_file)  # Already removes .z, preserves case
        
        if has_ip:
            ip = self.get_ip_address(z_file)
            slc = self.get_slc_name(ip, slc_programs)
            units = self.get_units_in_program(z_file)
        else:
            ip = slc = units = "N/A"
        
        return ComponentInfo(
            name=component_name,
            type=component_type,
            ip_address=ip,
            slc=slc,
            units=units,
            dependencies=dependencies
        )


class StyleManager:
    """Manages Excel styling for consistency."""
    
    def __init__(self):
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color=ExcelConfig.HEADER_COLOR,
            end_color=ExcelConfig.HEADER_COLOR,
            fill_type="solid"
        )
        self.header_alignment = Alignment(horizontal='left', vertical='center')
        
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    
    def apply_header_style(self, cell):
        """Apply header styling to a cell."""
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment


class WorksheetHelper:
    """Helper methods for worksheet operations."""
    
    @staticmethod
    def setup_headers(ws: Worksheet, headers: list[str], style_manager: StyleManager):
        """Set up worksheet headers with styling."""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            style_manager.apply_header_style(cell)
    
    @staticmethod
    def create_table(ws: Worksheet, table_name: str, ref: str):
        """Create an Excel Table with filtering enabled."""
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
    
    @staticmethod
    def auto_fit_columns(ws: Worksheet, max_width: int = 60):
        """Auto-fit column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 3, max_width)
            ws.column_dimensions[column_letter].width = adjusted_width


class ExcelGenerator:
    """Generates Excel spreadsheet with SattLine configuration."""
    
    def __init__(self, extractor: SattLineConfigExtractor):
        self.extractor = extractor
        self.config = ExcelConfig()
        self.workstation_mapper = WorkstationMapper()
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        # Create worksheets
        self.query_ws = self.wb.create_sheet("Query Tool", 0)
        self.dashboard_ws = self.wb.create_sheet("Dashboard")
        self.components_ws = self.wb.create_sheet("System Components")
        self.dependencies_ws = self.wb.create_sheet("Library Dependencies")
        
        # Initialize helpers
        self.style_manager = StyleManager()
    
    def generate(self, output_path: Path):
        """Generate complete Excel configuration sheet."""
        log.info("🚀 Starting Excel generation...")
        
        # Collect SLC programs
        slc_programs = self._collect_slc_programs()
        
        # Process all components
        component_data, dependency_data = self._process_all_components(slc_programs)
        
        # Store for later use in dependencies sheet
        self.all_component_data = component_data
        
        # Populate worksheets
        comp_row_count = self._populate_components_sheet(component_data)
        dep_row_count = self._populate_dependencies_sheet(dependency_data)
        
        # Create tables
        if comp_row_count > 1:
            WorksheetHelper.create_table(
                self.components_ws, "SystemComponents", f"A1:F{comp_row_count}"
            )
        
        if dep_row_count > 1:
            WorksheetHelper.create_table(
                self.dependencies_ws, "LibraryDependencies", f"A1:E{dep_row_count}"
            )
        
        # Create all sheets
        self._create_dashboard(comp_row_count, dep_row_count)
        self._create_station_configuration_sheet(component_data)
        self._create_configuration_summary_sheet()
        self._create_configuration_details_sheet()
        self._create_query_sheet(component_data)
        
        # Auto-fit columns
        for ws in [self.components_ws, self.dependencies_ws]:
            WorksheetHelper.auto_fit_columns(ws)
        
        # Save workbook
        self.wb.save(output_path)
        log.info(f"✅ Excel file saved to: {output_path}")
    
    def _collect_slc_programs(self) -> dict[str, str]:
        """Collect SLC programs for IP mapping (preserve original case)."""
        log.info("📡 Collecting SLC programs...")
        slc_programs = {}
        
        for z_file in self.extractor.get_z_files(self.extractor.unitlib_dir):
            file_stem = z_file.stem  # Preserve original case
            if "pbslc" in file_stem.lower() or "wd" in file_stem.lower():
                ip = self.extractor.get_ip_address(z_file)
                slc_programs[file_stem] = ip
        
        log.info(f"✓ Found {len(slc_programs)} SLC programs")
        return slc_programs
    
    def _process_all_components(self, slc_programs: dict[str, str]) -> tuple[list[ComponentInfo], list[tuple]]:
        """Process all components and their dependencies."""
        sections = [
            ("Program", self.extractor.unitlib_dir, True),
            ("Project Library", self.extractor.projectlib_dir, False),
            ("NNE Library", self.extractor.nnelib_dir, False),
            ("SG Library", self.extractor.sglib_dir, False),
        ]
        
        component_data = []
        dependency_data = []
        
        for component_type, directory, has_ip in sections:
            log.info(f"📦 Processing {component_type}s from {directory.name}...")
            z_files = self.extractor.get_z_files(directory)
            
            for z_file in z_files:
                comp_info = self.extractor.get_component_info(
                    z_file, component_type, has_ip, slc_programs
                )
                component_data.append(comp_info)
                
                # Add dependencies
                for dep in comp_info.dependencies:
                    dependency_data.append((comp_info.name, dep))
            
            log.info(f"✓ Processed {len(z_files)} {component_type}s")
        
        log.info(f"✓ Total: {len(component_data)} components, {len(dependency_data)} dependencies")
        return component_data, dependency_data
    
    def _populate_components_sheet(self, component_data: list[ComponentInfo]) -> int:
        """Populate System Components worksheet."""
        headers = ["ID", "Component_ID", "IP_Address", "SLC_Number", "Units_Served", "Component_Type"]
        WorksheetHelper.setup_headers(self.components_ws, headers, self.style_manager)
        
        for idx, comp in enumerate(component_data, start=2):
            self.components_ws.cell(row=idx, column=1, value=idx - 1)
            self.components_ws.cell(row=idx, column=2, value=comp.name)  # Original case preserved
            self.components_ws.cell(row=idx, column=3, value=comp.ip_address)
            self.components_ws.cell(row=idx, column=4, value=comp.slc)
            self.components_ws.cell(row=idx, column=5, value=comp.units)
            self.components_ws.cell(row=idx, column=6, value=comp.type)
        
        return len(component_data) + 1
    
    def _populate_dependencies_sheet(self, dependency_data: list[tuple]) -> int:
        """Populate Library Dependencies worksheet."""
        headers = ["Dependency_ID", "Component_ID", "Component_Type", "Library_Name", "Library_Type"]
        WorksheetHelper.setup_headers(self.dependencies_ws, headers, self.style_manager)
        
        # Build component type lookup for faster access (case-insensitive key, original name as value)
        component_type_map = {}
        for comp_data in self.all_component_data:
            component_type_map[comp_data.name.lower()] = comp_data.type
        
        for idx, (component, library) in enumerate(dependency_data, start=2):
            self.dependencies_ws.cell(row=idx, column=1, value=idx - 1)
            self.dependencies_ws.cell(row=idx, column=2, value=component)  # Original case preserved
            
            # Lookup component type (case-insensitive)
            component_type = component_type_map.get(component.lower(), "Unknown")
            self.dependencies_ws.cell(row=idx, column=3, value=component_type)
            
            self.dependencies_ws.cell(row=idx, column=4, value=library)  # Original case preserved
            
            # Lookup library type (case-insensitive)
            library_type = component_type_map.get(library.lower(), "Unknown")
            self.dependencies_ws.cell(row=idx, column=5, value=library_type)
        
        return len(dependency_data) + 1
    
    def _create_dashboard(self, component_row_count: int, dependency_row_count: int):
        """Create dashboard with KPIs."""
        ws = self.dashboard_ws
        
        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "SattLine Configuration Dashboard"
        title_cell.font = Font(bold=True, size=20, color=ExcelConfig.KPI_TITLE_COLOR)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        log.info("✓ Dashboard created")
    
    def _create_station_configuration_sheet(self, component_data: list[ComponentInfo]):
        """Create comprehensive station configuration sheet."""
        ws = self.wb.create_sheet("Station Configuration")
        
        # Title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = "🖥️ Workstation Configuration Overview"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Headers at row 3 - ADDED Physical_Location
        headers = [
            "Station_ID", "Station_Type", "Physical_Location", "Configuration_File", 
            "SLC_Number", "Programs", "Libraries", "Units_Served", "IP_Address"
        ]
        
        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            self.style_manager.apply_header_style(cell)
        
        # Build station configurations
        station_configs = self._build_station_configurations(component_data)
        
        current_row = header_row + 1
        for station_id in sorted(station_configs.keys()):
            config = station_configs[station_id]
            
            ws.cell(row=current_row, column=1, value=station_id)
            ws.cell(row=current_row, column=2, value=config['type'])
            ws.cell(row=current_row, column=3, value=self.workstation_mapper.get_physical_location(station_id))
            ws.cell(row=current_row, column=4, value=config['config_file'])  # Original case preserved
            
            # Aggregate SLC numbers from all programs
            slc_numbers = self._aggregate_slc_numbers(config, component_data)
            ws.cell(row=current_row, column=5, value=slc_numbers)
            
            # Programs list (original case preserved)
            ws.cell(row=current_row, column=6, value=", ".join(config['programs']) if config['programs'] else "No programs")
            
            # Libraries list (original case preserved)
            ws.cell(row=current_row, column=7, value=", ".join(config['libraries']) if config['libraries'] else "No libraries")
            
            # Aggregate units from all programs
            units_served = self._aggregate_units(config, component_data)
            ws.cell(row=current_row, column=8, value=units_served)
            
            # IP addresses
            ws.cell(row=current_row, column=9, value=config.get('ip_address', 'N/A'))
            
            current_row += 1
        
        # Create table
        if current_row > header_row + 1:
            table_ref = f"A{header_row}:I{current_row-1}"
            WorksheetHelper.create_table(ws, "StationConfiguration", table_ref)
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 50
        ws.column_dimensions['I'].width = 15
        
        log.info(f"✓ Station Configuration sheet created with {current_row - header_row - 1} stations")
    
    def _aggregate_slc_numbers(self, config: dict, component_data: list[ComponentInfo]) -> str:
        """Aggregate unique SLC numbers from all programs in the configuration (case-insensitive matching)."""
        slc_set = set()
        
        for program_name in config['programs']:
            # Case-insensitive comparison
            comp = next((c for c in component_data if c.name.lower() == program_name.lower()), None)
            if comp and comp.slc and comp.slc not in ["N/A", "No SLC"]:
                slc_set.add(comp.slc)
        
        if not slc_set:
            return "No SLC"
        
        return ", ".join(sorted(slc_set))
    
    def _aggregate_units(self, config: dict, component_data: list[ComponentInfo]) -> str:
        """Aggregate unique units from all programs in the configuration (case-insensitive matching)."""
        units_set = set()
        
        for program_name in config['programs']:
            # Case-insensitive comparison
            comp = next((c for c in component_data if c.name.lower() == program_name.lower()), None)
            if comp and comp.units:
                # Extract units from the format "(N) unit1, unit2, ..."
                units_text = comp.units
                if '(' in units_text and ')' in units_text:
                    parts = units_text.split(') ', 1)
                    if len(parts) == 2:
                        units_list = [u.strip() for u in parts[1].split(',')]
                        units_set.update(units_list)
        
        if not units_set:
            return "No units assigned"
        
        sorted_units = sorted(units_set)
        return f"({len(sorted_units)}) " + ", ".join(sorted_units)
    
    def _format_units_for_station(self, units_text: str) -> str:
        """Format units text for display."""
        if not units_text or units_text in ['N/A', 'No units assigned', 'No X-File', 'Error reading X-File']:
            return "No units assigned"
        
        if '(' in units_text and ')' in units_text:
            parts = units_text.split(') ', 1)
            if len(parts) == 2:
                return parts[1]
        
        return units_text
    
    def _build_station_configurations(self, component_data: list[ComponentInfo]) -> dict:
        """Build comprehensive configuration data for each station (case-insensitive matching)."""
        station_configs = {}
        
        # Initialize stations from mapper
        for config_file, workstations in self.workstation_mapper.workstation_map.items():
            for station_id in workstations:
                station_type = self._determine_station_type(station_id)
                
                if station_id not in station_configs:
                    station_configs[station_id] = {
                        'config_file': config_file,  # Original case preserved
                        'type': station_type,
                        'programs': [],
                        'libraries': [],
                        'units': None,
                        'ip_address': None,
                        'slc': None
                    }
        
        # Parse configuration files to get programs/libraries for each config
        configurations = self.extractor.parse_all_configuration_files()
        config_contents = {}
        
        for config in configurations:
            # Store with original case name, but create lowercase key for lookup
            config_contents[config.config_name.lower()] = {
                'programs': [p['name'] for p in config.programs],  # Original case preserved
                'libraries': [l['name'] for l in config.libraries]  # Original case preserved
            }
        
        # Populate station configs with programs and libraries from configuration files
        for station_id, config in station_configs.items():
            config_file_lower = config['config_file'].lower()
            
            if config_file_lower in config_contents:
                config['programs'] = config_contents[config_file_lower]['programs'].copy()
                config['libraries'] = config_contents[config_file_lower]['libraries'].copy()
                
                # Add transitive dependencies
                additional_libraries = set()
                for program_name in config['programs']:
                    # Case-insensitive comparison
                    comp = next((c for c in component_data if c.name.lower() == program_name.lower()), None)
                    if comp:
                        for dep in comp.dependencies:
                            # Case-insensitive check
                            dep_comp = next((c for c in component_data if c.name.lower() == dep.lower()), None)
                            if dep_comp and dep_comp.type != "Program":
                                additional_libraries.add(dep)  # Original case preserved
                
                for lib in additional_libraries:
                    # Case-insensitive check if library already exists
                    if not any(existing_lib.lower() == lib.lower() for existing_lib in config['libraries']):
                        config['libraries'].append(lib)
                
                config['libraries'].sort()
                config['programs'].sort()
        
        return station_configs
    
    def _determine_station_type(self, station_id: str) -> str:
        """Determine station type from station ID."""
        if station_id.startswith('LOP'):
            return "Local Operator Panel"
        elif station_id.startswith('OP'):
            return "Operator Station"
        elif station_id.startswith('OPC'):
            return "OPC Server"
        elif station_id.startswith('PRG'):
            return "Programmer Station"
        elif 'Journal' in station_id:
            return "Journal Server"
        else:
            return "Special System"
    
    def _create_configuration_summary_sheet(self):
        """Create summary worksheet showing overview of each configuration file."""
        ws = self.wb.create_sheet("Configuration Summary")
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "📊 Configuration Files Summary"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        configurations = self.extractor.parse_all_configuration_files()
        
        if not configurations:
            ws['A3'] = "No configuration files found or parsing failed"
            return
        
        headers = ["Configuration_Name", "Version", "Date", "Main_Program", "Total_Programs", "Total_Libraries"]
        header_row = 3
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            self.style_manager.apply_header_style(cell)
        
        current_row = header_row + 1
        for config in sorted(configurations, key=lambda x: x.config_name):
            ws.cell(row=current_row, column=1, value=config.config_name)  # Original case preserved
            ws.cell(row=current_row, column=2, value=config.version)
            ws.cell(row=current_row, column=3, value=config.date)
            ws.cell(row=current_row, column=4, value=config.main_program)  # Original case preserved
            ws.cell(row=current_row, column=5, value=len(config.programs))
            ws.cell(row=current_row, column=6, value=len(config.libraries))
            current_row += 1
        
        if current_row > header_row + 1:
            table_ref = f"A{header_row}:F{current_row-1}"
            WorksheetHelper.create_table(ws, "ConfigurationSummary", table_ref)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        log.info("✓ Configuration Summary sheet created")
    
    def _create_configuration_details_sheet(self):
        """Create detailed worksheet showing programs and libraries in each configuration file."""
        ws = self.wb.create_sheet("Configuration Details")
        
        # Title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = "📋 Configuration File Details"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        configurations = self.extractor.parse_all_configuration_files()
        
        if not configurations:
            ws['A3'] = "No configuration files found or parsing failed"
            return
        
        headers = [
            "Configuration_Name", "Component_Type", "Component_Name", "Directory",
            "Main_Program", "Version", "Date", "Total_Programs", "Total_Libraries"
        ]
        
        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            self.style_manager.apply_header_style(cell)
        
        current_row = header_row + 1
        
        for config in sorted(configurations, key=lambda x: x.config_name):
            # Add all programs
            for program in config.programs:
                ws.cell(row=current_row, column=1, value=config.config_name)  # Original case
                ws.cell(row=current_row, column=2, value="Program")
                ws.cell(row=current_row, column=3, value=program['name'])  # Original case
                ws.cell(row=current_row, column=4, value=program['directory'])
                ws.cell(row=current_row, column=5, value="Yes" if program['main_program'] else "No")
                ws.cell(row=current_row, column=6, value=config.version)
                ws.cell(row=current_row, column=7, value=config.date)
                ws.cell(row=current_row, column=8, value=len(config.programs))
                ws.cell(row=current_row, column=9, value=len(config.libraries))
                current_row += 1
            
            # Add all libraries
            for library in config.libraries:
                ws.cell(row=current_row, column=1, value=config.config_name)  # Original case
                ws.cell(row=current_row, column=2, value="Library")
                ws.cell(row=current_row, column=3, value=library['name'])  # Original case
                ws.cell(row=current_row, column=4, value=library['directory'])
                ws.cell(row=current_row, column=5, value="N/A")
                ws.cell(row=current_row, column=6, value=config.version)
                ws.cell(row=current_row, column=7, value=config.date)
                ws.cell(row=current_row, column=8, value=len(config.programs))
                ws.cell(row=current_row, column=9, value=len(config.libraries))
                current_row += 1
        
        if current_row > header_row + 1:
            table_ref = f"A{header_row}:I{current_row-1}"
            WorksheetHelper.create_table(ws, "ConfigurationDetails", table_ref)
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        
        log.info("✓ Configuration Details sheet created")
    
    def _create_query_sheet(self, component_data: list[ComponentInfo]):
        """Create impact analysis query tool."""
        ws = self.query_ws
        
        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "🎯 Change Impact Analysis Tool"
        title_cell.font = Font(bold=True, size=20, color=ExcelConfig.KPI_TITLE_COLOR)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35
        
        # Instructions
        ws.merge_cells('A3:H3')
        inst_cell = ws['A3']
        inst_cell.value = "Select programs/libraries you plan to change to see all impacted stations, units, dependencies, and SLCs"
        inst_cell.font = Font(italic=True, size=11, color="666666")
        inst_cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[3].height = 25
        
        # Selection area
        ws.merge_cells('A5:H5')
        sel_header = ws['A5']
        sel_header.value = "📝 SELECT COMPONENTS TO DOWNLOAD (up to 10)"
        sel_header.font = Font(bold=True, size=12, color="FFFFFF")
        sel_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        sel_header.alignment = Alignment(horizontal='center')
        
        # Create 10 selection dropdowns
        for i in range(10):
            row = 7 + i
            ws[f'A{row}'] = f"Component {i+1}:"
            ws[f'A{row}'].font = Font(bold=True, size=10)
            ws[f'A{row}'].alignment = Alignment(horizontal='right')
            
            ws.merge_cells(f'B{row}:D{row}')
            sel_cell = ws[f'B{row}']
            sel_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            sel_cell.border = self.style_manager.border
            
            # Data validation
            dv = DataValidation(
                type="list",
                formula1="'System Components'!$B$2:$B$1000",
                allow_blank=True
            )
            dv.add(sel_cell)
            ws.add_data_validation(dv)
            
            # Show component type
            ws[f'E{row}'] = f'=IF(B{row}="","",INDEX(\'System Components\'!F:F,MATCH(B{row},\'System Components\'!B:B,0)))'
            ws[f'E{row}'].font = Font(size=9, italic=True, color="666666")
        
        # Results section - Affected Workstations
        ws.merge_cells('A19:H19')
        results_header = ws['A19']
        results_header.value = "🖥️ AFFECTED WORKSTATIONS"
        results_header.font = Font(bold=True, size=12, color="FFFFFF")
        results_header.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        results_header.alignment = Alignment(horizontal='center')
        
        # Build reverse mapping: component -> workstations (use original case for display)
        component_to_workstations = {}
        component_original_case = {}  # Map lowercase to original case
        
        for config_file, workstations in self.workstation_mapper.workstation_map.items():
            configurations = self.extractor.parse_all_configuration_files()
            for config in configurations:
                if config.config_name.lower() == config_file.lower():
                    # Map all programs in this config to these workstations
                    for program in config.programs:
                        prog_name = program['name']
                        prog_name_lower = prog_name.lower()
                        component_original_case[prog_name_lower] = prog_name  # Store original case
                        
                        if prog_name_lower not in component_to_workstations:
                            component_to_workstations[prog_name_lower] = set()
                        component_to_workstations[prog_name_lower].update(workstations)
                    
                    # Map all libraries in this config to these workstations
                    for library in config.libraries:
                        lib_name = library['name']
                        lib_name_lower = lib_name.lower()
                        component_original_case[lib_name_lower] = lib_name  # Store original case
                        
                        if lib_name_lower not in component_to_workstations:
                            component_to_workstations[lib_name_lower] = set()
                        component_to_workstations[lib_name_lower].update(workstations)
        
        # Create formulas to display affected workstations
        ws['A21'] = "Workstations:"
        ws['A21'].font = Font(bold=True, size=10)
        ws['A21'].alignment = Alignment(horizontal='right')
        
        # Create a lookup table (with original case for display)
        ws['J1'] = "Component_Lookup"
        ws['K1'] = "Workstations"
        ws['J1'].font = Font(bold=True)
        ws['K1'].font = Font(bold=True)
        
        lookup_row = 2
        for comp_name_lower, stations in sorted(component_to_workstations.items()):
            # Use original case for display in lookup table
            original_name = component_original_case.get(comp_name_lower, comp_name_lower)
            ws[f'J{lookup_row}'] = original_name
            ws[f'K{lookup_row}'] = ", ".join(sorted(stations))
            lookup_row += 1
        
        # Create formula that concatenates unique workstations using TEXTJOIN and VLOOKUPs
        vlookup_formulas = []
        for i in range(10):
            row_ref = 7 + i
            # Case-insensitive VLOOKUP - match against exact values in lookup table
            vlookup_formulas.append(f'IFERROR(VLOOKUP(B{row_ref},J:K,2,FALSE),"")')
        
        # Join all workstations with comma
        ws.merge_cells('B21:H21')
        ws['B21'] = f'=TEXTJOIN(", ",TRUE,{",".join(vlookup_formulas)})'
        ws['B21'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Hide the lookup columns
        ws.column_dimensions['J'].hidden = True
        ws.column_dimensions['K'].hidden = True
        
        # Set column widths for visible columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        log.info("✓ Query Tool created with workstation impact analysis")


def main():
    """Main entry point."""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    try:
        root_dir = Path(r"C:\Users\SQHJ\OneDrive - Novo Nordisk\Workspace\Libs\GC")
        output_file = Path("SattLine_Configuration.xlsx")
        
        extractor = SattLineConfigExtractor(root_dir)
        generator = ExcelGenerator(extractor)
        generator.generate(output_file)
        
        print(f"✅ Configuration Excel file generated successfully: {output_file}")
    
    except Exception as e:
        log.error(f"Failed to generate Excel file: {e}", exc_info=True)
        print(f"❌ Error: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())