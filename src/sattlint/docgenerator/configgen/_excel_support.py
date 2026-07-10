from __future__ import annotations

from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from ._types import ExcelConfig


class StyleManager:
    def __init__(self):
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color=ExcelConfig.HEADER_COLOR,
            end_color=ExcelConfig.HEADER_COLOR,
            fill_type="solid",
        )
        self.header_alignment = Alignment(horizontal="left", vertical="center")
        self.border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

    def apply_header_style(self, cell: Cell) -> None:
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment


class WorksheetHelper:
    @staticmethod
    def setup_headers(ws: Worksheet, headers: list[str], style_manager: StyleManager):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            style_manager.apply_header_style(cell)

    @staticmethod
    def create_table(ws: Worksheet, table_name: str, ref: str):
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    @staticmethod
    def auto_fit_columns(ws: Worksheet, max_width: int = 60):
        for idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(idx)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 3, max_width)
