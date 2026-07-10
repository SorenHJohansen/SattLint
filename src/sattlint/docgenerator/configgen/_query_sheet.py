from __future__ import annotations

from typing import Any, cast

from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill

from ._types import ComponentInfo, ExcelConfig, _WorksheetCellTarget


def create_query_sheet(generator: Any, component_data: list[ComponentInfo]) -> None:  # noqa: PLR0915
    ws = generator.query_ws
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "🎯 Change Impact Analysis Tool"
    title_cell.font = Font(bold=True, size=20, color=ExcelConfig.KPI_TITLE_COLOR)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A3:H3")
    inst_cell = ws["A3"]
    inst_cell.value = (
        "Select programs/libraries you plan to change to see all impacted stations, units, dependencies, and SLCs"
    )
    inst_cell.font = Font(italic=True, size=11, color="666666")
    inst_cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[3].height = 25

    ws.merge_cells("A5:H5")
    sel_header = ws["A5"]
    sel_header.value = "📝 SELECT COMPONENTS TO DOWNLOAD (up to 10)"
    sel_header.font = Font(bold=True, size=12, color="FFFFFF")
    sel_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    sel_header.alignment = Alignment(horizontal="center")

    for index in range(10):
        row = 7 + index
        ws[f"A{row}"] = f"Component {index + 1}:"
        ws[f"A{row}"].font = Font(bold=True, size=10)
        ws[f"A{row}"].alignment = Alignment(horizontal="right")
        ws.merge_cells(f"B{row}:D{row}")
        sel_cell = cast(Cell, ws[f"B{row}"])
        sel_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        sel_cell.border = generator.style_manager.border
        dv = generator._component_validation()
        cast(_WorksheetCellTarget, dv).add(sel_cell)
        ws.add_data_validation(dv)
        ws[f"E{row}"] = f"=IF(B{row}=\"\",\"\",INDEX('System Components'!F:F,MATCH(B{row},'System Components'!B:B,0)))"
        ws[f"E{row}"].font = Font(size=9, italic=True, color="666666")

    ws.merge_cells("A19:H19")
    results_header = ws["A19"]
    results_header.value = "AFFECTED WORKSTATIONS"
    results_header.font = Font(bold=True, size=12, color="FFFFFF")
    results_header.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    results_header.alignment = Alignment(horizontal="center")

    component_to_workstations: dict[str, set[str]] = {}
    component_original_case: dict[str, str] = {}
    configurations = generator.extractor.parse_all_configuration_files()
    for config_file, workstations in generator.workstation_mapper.workstation_map.items():
        for config in configurations:
            if config.config_name.lower() != config_file.lower():
                continue
            for program in config.programs:
                prog_name = program["name"]
                prog_key = prog_name.lower()
                component_original_case[prog_key] = prog_name
                component_to_workstations.setdefault(prog_key, set()).update(workstations)
            for library in config.libraries:
                lib_name = library["name"]
                lib_key = lib_name.lower()
                component_original_case[lib_key] = lib_name
                component_to_workstations.setdefault(lib_key, set()).update(workstations)

    ws["A21"] = "Workstations:"
    ws["A21"].font = Font(bold=True, size=10)
    ws["A21"].alignment = Alignment(horizontal="right")
    ws["J1"] = "Component_Lookup"
    ws["K1"] = "Workstations"
    ws["J1"].font = Font(bold=True)
    ws["K1"].font = Font(bold=True)

    lookup_row = 2
    for comp_name_lower, stations in sorted(component_to_workstations.items()):
        ws[f"J{lookup_row}"] = component_original_case.get(comp_name_lower, comp_name_lower)
        ws[f"K{lookup_row}"] = ", ".join(sorted(stations))
        lookup_row += 1

    formulas = [f'IFERROR(VLOOKUP(B{7 + index},J:K,2,FALSE),"")' for index in range(10)]
    ws.merge_cells("B21:H21")
    ws["B21"] = f'=TEXTJOIN(", ",TRUE,{",".join(formulas)})'
    ws["B21"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["J"].hidden = True
    ws.column_dimensions["K"].hidden = True
    for column in "ABCDEFGH":
        ws.column_dimensions[column].width = 15 if column in {"A", "F", "G", "H"} else 20
