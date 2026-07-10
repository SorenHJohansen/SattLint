from __future__ import annotations

import logging
from collections.abc import Mapping
from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from ._excel_support import StyleManager, WorksheetHelper
from ._extraction import SattLineConfigExtractor, WorkstationMapper
from ._query_sheet import create_query_sheet
from ._types import ComponentInfo, ConfigurationContents, DependencyRow, ExcelConfig, StationConfiguration

log = logging.getLogger("SattLint")


class ExcelGenerator:
    def __init__(self, extractor: SattLineConfigExtractor):
        self.extractor = extractor
        self.config = ExcelConfig()
        self.workstation_mapper = WorkstationMapper()
        self.wb = Workbook()
        if self.wb.active is not None:
            self.wb.remove(self.wb.active)
        self.query_ws = self.wb.create_sheet("Query Tool", 0)
        self.dashboard_ws = self.wb.create_sheet("Dashboard")
        self.components_ws = self.wb.create_sheet("System Components")
        self.dependencies_ws = self.wb.create_sheet("Library Dependencies")
        self.style_manager = StyleManager()

    def generate(self, output_path: Path):
        log.info("🚀 Starting Excel generation...")
        slc_programs = self._collect_slc_programs()
        component_data, dependency_data = self._process_all_components(slc_programs)
        self.all_component_data = component_data
        comp_row_count = self._populate_components_sheet(component_data)
        dep_row_count = self._populate_dependencies_sheet(dependency_data)
        if comp_row_count > 1:
            WorksheetHelper.create_table(self.components_ws, "SystemComponents", f"A1:F{comp_row_count}")
        if dep_row_count > 1:
            WorksheetHelper.create_table(self.dependencies_ws, "LibraryDependencies", f"A1:E{dep_row_count}")
        self._create_dashboard(comp_row_count, dep_row_count)
        self._create_station_configuration_sheet(component_data)
        self._create_configuration_summary_sheet()
        self._create_configuration_details_sheet()
        self._create_query_sheet(component_data)
        for worksheet in [self.components_ws, self.dependencies_ws]:
            WorksheetHelper.auto_fit_columns(worksheet)
        self.wb.save(output_path)
        log.info("✅ Excel file saved to: %s", output_path)

    def _component_validation(self) -> DataValidation:
        return DataValidation(type="list", formula1="'System Components'!$B$2:$B$1000", allow_blank=True)

    def _collect_slc_programs(self) -> dict[str, str]:
        log.info("📡 Collecting SLC programs...")
        slc_programs: dict[str, str] = {}
        for z_file in self.extractor.get_z_files(self.extractor.unitlib_dir):
            file_stem = z_file.stem
            if "pbslc" in file_stem.lower() or "wd" in file_stem.lower():
                slc_programs[file_stem] = self.extractor.get_ip_address(z_file)
        log.info("✓ Found %d SLC programs", len(slc_programs))
        return slc_programs

    def _process_all_components(self, slc_programs: dict[str, str]) -> tuple[list[ComponentInfo], list[DependencyRow]]:
        sections = [
            ("Program", self.extractor.unitlib_dir, True),
            ("Project Library", self.extractor.projectlib_dir, False),
            ("NNE Library", self.extractor.nnelib_dir, False),
            ("SG Library", self.extractor.sglib_dir, False),
        ]
        component_data: list[ComponentInfo] = []
        dependency_data: list[DependencyRow] = []
        for component_type, directory, has_ip in sections:
            log.info("📦 Processing %ss from %s...", component_type, directory.name)
            z_files = self.extractor.get_z_files(directory)
            for z_file in z_files:
                comp_info = self.extractor.get_component_info(z_file, component_type, has_ip, slc_programs)
                component_data.append(comp_info)
                for dep in comp_info.dependencies:
                    dependency_data.append((comp_info.name, dep))
            log.info("✓ Processed %d %ss", len(z_files), component_type)
        log.info("✓ Total: %d components, %d dependencies", len(component_data), len(dependency_data))
        return component_data, dependency_data

    def _populate_components_sheet(self, component_data: list[ComponentInfo]) -> int:
        WorksheetHelper.setup_headers(
            self.components_ws,
            ["ID", "Component_ID", "IP_Address", "SLC_Number", "Units_Served", "Component_Type"],
            self.style_manager,
        )
        for idx, comp in enumerate(component_data, start=2):
            self.components_ws.cell(row=idx, column=1, value=idx - 1)
            self.components_ws.cell(row=idx, column=2, value=comp.name)
            self.components_ws.cell(row=idx, column=3, value=comp.ip_address)
            self.components_ws.cell(row=idx, column=4, value=comp.slc)
            self.components_ws.cell(row=idx, column=5, value=comp.units)
            self.components_ws.cell(row=idx, column=6, value=comp.type)
        return len(component_data) + 1

    def _populate_dependencies_sheet(self, dependency_data: list[DependencyRow]) -> int:
        WorksheetHelper.setup_headers(
            self.dependencies_ws,
            ["Dependency_ID", "Component_ID", "Component_Type", "Library_Name", "Library_Type"],
            self.style_manager,
        )
        component_type_map = {component.name.lower(): component.type for component in self.all_component_data}
        for idx, (component, library) in enumerate(dependency_data, start=2):
            self.dependencies_ws.cell(row=idx, column=1, value=idx - 1)
            self.dependencies_ws.cell(row=idx, column=2, value=component)
            self.dependencies_ws.cell(row=idx, column=3, value=component_type_map.get(component.lower(), "Unknown"))
            self.dependencies_ws.cell(row=idx, column=4, value=library)
            self.dependencies_ws.cell(row=idx, column=5, value=component_type_map.get(library.lower(), "Unknown"))
        return len(dependency_data) + 1

    def _create_dashboard(self, _component_row_count: int, _dependency_row_count: int):
        ws = self.dashboard_ws
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = "SattLine Configuration Dashboard"
        title_cell.font = Font(bold=True, size=20, color=ExcelConfig.KPI_TITLE_COLOR)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        log.info("✓ Dashboard created")

    def _create_station_configuration_sheet(self, component_data: list[ComponentInfo]):
        ws = self.wb.create_sheet("Station Configuration")
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = "Workstation Configuration Overview"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        headers = [
            "Station_ID",
            "Station_Type",
            "Physical_Location",
            "Configuration_File",
            "SLC_Number",
            "Programs",
            "Libraries",
            "Units_Served",
            "IP_Address",
        ]
        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            self.style_manager.apply_header_style(cell)
        station_configs = self._build_station_configurations(component_data)
        current_row = header_row + 1
        for station_id in sorted(station_configs.keys()):
            config = station_configs[station_id]
            ws.cell(row=current_row, column=1, value=station_id)
            ws.cell(row=current_row, column=2, value=config["type"])
            ws.cell(row=current_row, column=3, value=self.workstation_mapper.get_physical_location(station_id))
            ws.cell(row=current_row, column=4, value=config["config_file"])
            ws.cell(row=current_row, column=5, value=self._aggregate_slc_numbers(config, component_data))
            ws.cell(
                row=current_row, column=6, value=", ".join(config["programs"]) if config["programs"] else "No programs"
            )
            ws.cell(
                row=current_row,
                column=7,
                value=", ".join(config["libraries"]) if config["libraries"] else "No libraries",
            )
            ws.cell(row=current_row, column=8, value=self._aggregate_units(config, component_data))
            ws.cell(row=current_row, column=9, value=config.get("ip_address") or "N/A")
            current_row += 1
        if current_row > header_row + 1:
            WorksheetHelper.create_table(ws, "StationConfiguration", f"A{header_row}:I{current_row - 1}")
        for column, width in {"A": 20, "B": 25, "C": 25, "D": 25, "E": 25, "F": 40, "G": 40, "H": 50, "I": 15}.items():
            ws.column_dimensions[column].width = width
        log.info("✓ Station Configuration sheet created with %d stations", current_row - header_row - 1)

    def _aggregate_slc_numbers(self, config: Mapping[str, object], component_data: list[ComponentInfo]) -> str:
        slc_set: set[str] = set()
        for program_name in cast(list[str], config["programs"]):
            comp = next(
                (component for component in component_data if component.name.lower() == program_name.lower()), None
            )
            if comp and comp.slc and comp.slc not in ["N/A", "No SLC"]:
                slc_set.add(comp.slc)
        return "No SLC" if not slc_set else ", ".join(sorted(slc_set))

    def _aggregate_units(self, config: Mapping[str, object], component_data: list[ComponentInfo]) -> str:
        units_set: set[str] = set()
        for program_name in cast(list[str], config["programs"]):
            comp = next(
                (component for component in component_data if component.name.lower() == program_name.lower()), None
            )
            if comp and comp.units and "(" in comp.units and ")" in comp.units:
                parts = comp.units.split(") ", 1)
                if len(parts) == 2:
                    units_set.update(unit.strip() for unit in parts[1].split(","))
        if not units_set:
            return "No units assigned"
        sorted_units = sorted(units_set)
        return f"({len(sorted_units)}) " + ", ".join(sorted_units)

    def _format_units_for_station(self, units_text: str) -> str:
        if not units_text or units_text in ["N/A", "No units assigned", "No X-File", "Error reading X-File"]:
            return "No units assigned"
        if "(" in units_text and ")" in units_text:
            parts = units_text.split(") ", 1)
            if len(parts) == 2:
                return parts[1]
        return units_text

    def _build_station_configurations(self, component_data: list[ComponentInfo]) -> dict[str, StationConfiguration]:
        station_configs: dict[str, StationConfiguration] = {}

        for config_file, workstations in self.workstation_mapper.workstation_map.items():
            for station_id in workstations:
                if station_id not in station_configs:
                    station_configs[station_id] = {
                        "config_file": config_file,
                        "type": self._determine_station_type(station_id),
                        "programs": [],
                        "libraries": [],
                        "units": None,
                        "ip_address": None,
                        "slc": None,
                    }

        configurations = self.extractor.parse_all_configuration_files()
        config_contents: dict[str, ConfigurationContents] = {}
        for config in configurations:
            config_contents[config.config_name.lower()] = {
                "programs": [program["name"] for program in config.programs],
                "libraries": [library["name"] for library in config.libraries],
            }

        for station_config in station_configs.values():
            config_file_lower = station_config["config_file"].lower()
            if config_file_lower not in config_contents:
                continue

            station_config["programs"] = config_contents[config_file_lower]["programs"].copy()
            station_config["libraries"] = config_contents[config_file_lower]["libraries"].copy()

            additional_libraries: set[str] = set()
            for program_name in station_config["programs"]:
                comp = next((c for c in component_data if c.name.lower() == program_name.lower()), None)
                if comp is None:
                    continue
                for dependency in comp.dependencies:
                    dep_comp = next((c for c in component_data if c.name.lower() == dependency.lower()), None)
                    if dep_comp and dep_comp.type != "Program":
                        additional_libraries.add(dependency)

            for library in additional_libraries:
                if not any(existing.lower() == library.lower() for existing in station_config["libraries"]):
                    station_config["libraries"].append(library)

            station_config["libraries"].sort()
            station_config["programs"].sort()

        return station_configs

    def _determine_station_type(self, station_id: str) -> str:
        if station_id.startswith("LOP"):
            return "Local Operator Panel"
        if station_id.startswith("OPC"):
            return "OPC Server"
        if station_id.startswith("OP"):
            return "Operator Station"
        if station_id.startswith("PRG"):
            return "Programmer Station"
        if "Journal" in station_id:
            return "Journal Server"
        return "Special System"

    def _create_configuration_summary_sheet(self):
        ws = self.wb.create_sheet("Configuration Summary")
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "📊 Configuration Files Summary"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        configurations = self.extractor.parse_all_configuration_files()
        if not configurations:
            ws["A3"] = "No configuration files found or parsing failed"
            return
        header_row = 3
        for col, header in enumerate(
            ["Configuration_Name", "Version", "Date", "Main_Program", "Total_Programs", "Total_Libraries"], start=1
        ):
            self.style_manager.apply_header_style(ws.cell(row=header_row, column=col, value=header))
        current_row = header_row + 1
        for config in sorted(configurations, key=lambda value: value.config_name):
            ws.cell(row=current_row, column=1, value=config.config_name)
            ws.cell(row=current_row, column=2, value=config.version)
            ws.cell(row=current_row, column=3, value=config.date)
            ws.cell(row=current_row, column=4, value=config.main_program)
            ws.cell(row=current_row, column=5, value=len(config.programs))
            ws.cell(row=current_row, column=6, value=len(config.libraries))
            current_row += 1
        if current_row > header_row + 1:
            WorksheetHelper.create_table(ws, "ConfigurationSummary", f"A{header_row}:F{current_row - 1}")
        for column, width in {"A": 25, "B": 12, "C": 25, "D": 25, "E": 15, "F": 15}.items():
            ws.column_dimensions[column].width = width
        log.info("✓ Configuration Summary sheet created")

    def _create_configuration_details_sheet(self):
        ws = self.wb.create_sheet("Configuration Details")
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = "📋 Configuration File Details"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        configurations = self.extractor.parse_all_configuration_files()
        if not configurations:
            ws["A3"] = "No configuration files found or parsing failed"
            return
        headers = [
            "Configuration_Name",
            "Component_Type",
            "Component_Name",
            "Directory",
            "Main_Program",
            "Version",
            "Date",
            "Total_Programs",
            "Total_Libraries",
        ]
        header_row = 3
        for col, header in enumerate(headers, start=1):
            self.style_manager.apply_header_style(ws.cell(row=header_row, column=col, value=header))
        current_row = header_row + 1
        for config in sorted(configurations, key=lambda value: value.config_name):
            for program in config.programs:
                ws.cell(row=current_row, column=1, value=config.config_name)
                ws.cell(row=current_row, column=2, value="Program")
                ws.cell(row=current_row, column=3, value=program["name"])
                ws.cell(row=current_row, column=4, value=program["directory"])
                ws.cell(row=current_row, column=5, value="Yes" if program["main_program"] else "No")
                ws.cell(row=current_row, column=6, value=config.version)
                ws.cell(row=current_row, column=7, value=config.date)
                ws.cell(row=current_row, column=8, value=len(config.programs))
                ws.cell(row=current_row, column=9, value=len(config.libraries))
                current_row += 1
            for library in config.libraries:
                ws.cell(row=current_row, column=1, value=config.config_name)
                ws.cell(row=current_row, column=2, value="Library")
                ws.cell(row=current_row, column=3, value=library["name"])
                ws.cell(row=current_row, column=4, value=library["directory"])
                ws.cell(row=current_row, column=5, value="N/A")
                ws.cell(row=current_row, column=6, value=config.version)
                ws.cell(row=current_row, column=7, value=config.date)
                ws.cell(row=current_row, column=8, value=len(config.programs))
                ws.cell(row=current_row, column=9, value=len(config.libraries))
                current_row += 1
        if current_row > header_row + 1:
            WorksheetHelper.create_table(ws, "ConfigurationDetails", f"A{header_row}:I{current_row - 1}")
        for column, width in {"A": 25, "B": 15, "C": 30, "D": 40, "E": 12, "F": 12, "G": 25, "H": 15, "I": 15}.items():
            ws.column_dimensions[column].width = width
        log.info("✓ Configuration Details sheet created")

    def _create_query_sheet(self, component_data: list[ComponentInfo]):
        create_query_sheet(self, component_data)
